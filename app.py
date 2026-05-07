"""
LP Training - Generador de Informes Mensuales
App web para procesar CSV de TrainingPeaks y rellenar la plantilla.
"""
import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
import shutil
import io
import os
import tempfile

# Tipos que NO cuentan como entreno programado
EXCLUDE_FROM_ADHERENCE = {'Day Off', 'Note'}
RUN_TYPE = 'Run'

PLANTILLA_PATH = 'Plantilla_Informe_LP_Training.xlsx'

# ==========================================================
# LÓGICA DE PROCESAMIENTO (igual que el script CLI)
# ==========================================================

def cargar_csv(file_obj):
    df = pd.read_csv(file_obj)
    df['WorkoutDay'] = pd.to_datetime(df['WorkoutDay'])
    for col in ['DistanceInMeters', 'TimeTotalInHours', 'PlannedDuration',
                'PlannedDistanceInMeters', 'VelocityAverage']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    return df


def detectar_mes(df):
    """
    Detecta el mes objetivo del CSV. Como ahora se exporta con rango ampliado
    (incluye días del mes anterior y/o siguiente), tomamos el mes que más
    fechas tiene en el CSV.
    """
    fechas = df['WorkoutDay']
    meses = fechas.dt.to_period('M').value_counts()
    mes_objetivo = meses.idxmax()
    return mes_objetivo.year, mes_objetivo.month


def filtrar_df_al_mes(df, year, month):
    """
    Filtra el dataframe a los días naturales del mes (1 al último día).
    Esto es lo que usamos para los KPIs globales: KM totales, horas, adherencia, etc.
    """
    primer_dia = datetime(year, month, 1)
    if month == 12:
        ultimo_dia = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        ultimo_dia = datetime(year, month + 1, 1) - timedelta(days=1)

    mask = (df['WorkoutDay'] >= primer_dia) & (df['WorkoutDay'] <= ultimo_dia)
    return df[mask]


def calcular_adherencia(df):
    programados_mask = ~df['WorkoutType'].isin(EXCLUDE_FROM_ADHERENCE)
    programados = programados_mask.sum()
    completados = (programados_mask & (df['TimeTotalInHours'] > 0)).sum()
    return int(programados), int(completados)


def calcular_metricas_running(df):
    runs = df[df['WorkoutType'] == RUN_TYPE]
    km_totales = runs['DistanceInMeters'].sum() / 1000
    rodaje_max = runs['DistanceInMeters'].max() / 1000 if len(runs) else 0

    runs_con_datos = runs[runs['DistanceInMeters'] > 0]
    if len(runs_con_datos) > 0:
        tiempo_total_seg = runs_con_datos['TimeTotalInHours'].sum() * 3600
        dist_total_km = runs_con_datos['DistanceInMeters'].sum() / 1000
        if dist_total_km > 0:
            seg_por_km = tiempo_total_seg / dist_total_km
            mins = int(seg_por_km // 60)
            segs = int(seg_por_km % 60)
            ritmo_str = f"{mins:02d}:{segs:02d}"
        else:
            ritmo_str = "00:00"
    else:
        ritmo_str = "00:00"

    return round(km_totales, 1), round(rodaje_max, 1), ritmo_str


def calcular_horas_totales(df):
    total_horas = df['TimeTotalInHours'].sum()
    h = int(total_horas)
    m = int((total_horas - h) * 60)
    s = int(((total_horas - h) * 60 - m) * 60)
    return f"{h}:{m:02d}:{s:02d}"


def calcular_por_semana(df, year, month):
    """
    Agrupa km de carrera y horas por semanas naturales LUN-DOM.
    Cada semana se asigna al mes en el que cae su LUNES.

    Para procesar el mes 'month':
    - Encontrar el primer lunes cuyo mes sea 'month' (puede ser día 1 si es lunes,
      o anterior si el día 1 no es lunes pero pertenece a una semana cuyo lunes
      ya está en el mes anterior — en ese caso, esa semana es del mes anterior).
    - Generar 4-5 semanas lun-dom hasta que el lunes salga del mes.

    Devuelve también un dict con info de qué días de datos faltan (para avisos).
    """
    primer_dia = datetime(year, month, 1)
    if month == 12:
        ultimo_dia = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        ultimo_dia = datetime(year, month + 1, 1) - timedelta(days=1)

    # Regla "DÍA 1 MANDA + Días del mes":
    # - Una semana se asigna a este mes si contiene al menos 1 día del mes.
    # - Los KM/horas de cada semana cuentan solo los DÍAS que pertenecen al mes
    #   (no los días que están en el mes anterior o siguiente).
    #
    # Esto significa:
    # - S1 = la semana lun-dom que contiene el día 1 del mes
    # - Última semana = la última semana lun-dom que contiene algún día del mes
    # - Las semanas frontera aparecen en dos meses pero cada informe muestra
    #   solo sus días correspondientes.

    # Lunes de la S1: lunes de la semana que contiene el día 1
    if primer_dia.weekday() == 0:
        primer_lunes = primer_dia
    else:
        primer_lunes = primer_dia - timedelta(days=primer_dia.weekday())

    # Construir las semanas: empezar en la S1 y avanzar mientras la semana
    # actual tenga al menos 1 día del mes.
    rangos = []
    lunes = primer_lunes
    while True:
        domingo = lunes + timedelta(days=6)
        # ¿Esta semana contiene algún día del mes?
        # La semana contiene días del mes si: lunes <= último_día_del_mes Y domingo >= primer_día
        if lunes > ultimo_dia:
            break  # La semana ya está completamente fuera del mes (después)
        if domingo < primer_dia:
            # La semana está completamente antes del mes — avanzar
            lunes = lunes + timedelta(days=7)
            continue
        rangos.append((lunes, domingo))
        lunes = lunes + timedelta(days=7)
        if len(rangos) > 10:  # protección
            break

    # Calcular métricas por semana
    semanas = []
    dias_faltantes = []

    # Determinar el rango real cubierto por el CSV (para detectar huecos en bordes)
    if len(df) > 0:
        primer_dia_csv = df['WorkoutDay'].min().date()
        ultimo_dia_csv = df['WorkoutDay'].max().date()
    else:
        primer_dia_csv = None
        ultimo_dia_csv = None

    for n, (ini, fin) in enumerate(rangos, start=1):
        # Recortar el rango de la semana al rango del mes (solo días del mes)
        ini_recortado = max(ini, primer_dia)
        fin_recortado = min(fin, ultimo_dia)

        # Filtrar el df por los días recortados (los que sí están en el mes)
        mask = (df['WorkoutDay'] >= ini_recortado) & (df['WorkoutDay'] <= fin_recortado)
        df_sem = df[mask]

        km_sem = df_sem[df_sem['WorkoutType'] == RUN_TYPE]['DistanceInMeters'].sum() / 1000
        horas_sem = df_sem['TimeTotalInHours'].sum()
        h = int(horas_sem)
        m = int((horas_sem - h) * 60)
        s = int(((horas_sem - h) * 60 - m) * 60)
        horas_str = f"{h}:{m:02d}:{s:02d}"

        # La etiqueta muestra el rango lun-dom completo (visualmente claro),
        # aunque solo se cuenten los días del mes.
        if ini.month == fin.month:
            label = f"S {n} ({ini.day} a {fin.day})"
        else:
            label = f"S {n} ({ini.day}/{ini.month} a {fin.day}/{fin.month})"

        semanas.append({
            'label': label,
            'km': round(km_sem, 1),
            'horas': horas_str,
            'horas_fraccion_dia': horas_sem / 24.0,  # Para gráficos: fracción de día (0-1)
            'inicio': ini_recortado,  # solo días del mes
            'fin': fin_recortado,
        })

        # Solo avisar de días faltantes si están FUERA del rango cubierto por el CSV
        if primer_dia_csv is not None:
            dia_actual = ini_recortado
            while dia_actual <= fin_recortado:
                if dia_actual.date() < primer_dia_csv or dia_actual.date() > ultimo_dia_csv:
                    dias_faltantes.append(dia_actual.date())
                dia_actual += timedelta(days=1)

    # Rellenar hasta 5 huecos
    semanas = semanas[:5]
    while len(semanas) < 5:
        semanas.append({'label': f'S {len(semanas)+1} (-)', 'km': 0, 'horas': '0:00:00',
                         'horas_fraccion_dia': 0.0, 'inicio': None, 'fin': None})

    return semanas, dias_faltantes


def nombre_mes(month):
    nombres = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
               'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
    return nombres[month - 1]


def formatear_timedelta(td):
    """Convierte timedelta a string HH:MM:SS."""
    if td is None:
        return None
    if isinstance(td, str):
        return td
    if hasattr(td, 'total_seconds'):
        total = int(td.total_seconds())
        h = total // 3600
        m = (total % 3600) // 60
        s = total % 60
        return f"{h}:{m:02d}:{s:02d}"
    if hasattr(td, 'hour'):  # datetime.time
        return f"{td.hour}:{td.minute:02d}:{td.second:02d}"
    return str(td)


def formatear_ritmo(t):
    """Convierte time a string MM:SS (ritmo min/km)."""
    if t is None:
        return None
    if isinstance(t, str):
        return t
    if hasattr(t, 'minute'):
        return f"{t.minute:02d}:{t.second:02d}"
    return str(t)


def extraer_mes_anterior(file_obj):
    """
    Detecta automáticamente el formato del Excel del mes anterior y extrae los datos.
    Soporta dos formatos:
    - Tipo A: Excel "técnico" antiguo con hoja 'INFORME ' y datos en G23/G25/G27/G29
    - Tipo B: Excel generado por esta misma web con hoja 'INPUTS' y datos en C12/C13/C19/F13
    Devuelve dict con 4 valores + fecha_inicio_plan, o None si no se puede leer.
    """
    try:
        wb = load_workbook(file_obj, data_only=True)
        hojas = wb.sheetnames

        # Tipo B: Excel generado por la web
        if 'INPUTS' in hojas:
            ws = wb['INPUTS']
            return {
                'distancia': ws['C12'].value,
                'tiempo': formatear_timedelta(ws['C13'].value),
                'ritmo': formatear_ritmo(ws['C19'].value) if ws['C19'].value else None,
                'rodaje': ws['F13'].value,
                'fecha_inicio_plan': formatear_fecha(ws['C8'].value),  # Fecha de inicio del plan
                'distancia_objetivo': ws['C7'].value,  # Distancia objetivo (también la copiamos)
                'formato': 'web',
            }

        # Tipo A: Excel técnico antiguo (hoja "INFORME " con espacio o sin él)
        hoja_informe = None
        for nombre in hojas:
            if 'INFORME' in nombre.upper():
                hoja_informe = nombre
                break

        if hoja_informe:
            ws = wb[hoja_informe]
            return {
                'distancia': ws['G23'].value,
                'tiempo': formatear_timedelta(ws['G25'].value),
                'ritmo': formatear_ritmo(ws['G27'].value),
                'rodaje': ws['G29'].value,
                'fecha_inicio_plan': formatear_fecha(ws['D6'].value),  # Fecha en D6
                'distancia_objetivo': None,  # No se extrae automáticamente del Excel técnico
                'formato': 'tecnico',
            }

        return None
    except Exception as e:
        return None


def formatear_fecha(d):
    """Convierte datetime a string DD/MM/YYYY."""
    if d is None:
        return None
    if isinstance(d, str):
        return d
    if hasattr(d, 'strftime'):
        return d.strftime('%d/%m/%Y')
    return str(d)


def rellenar_plantilla(plantilla_path, datos, datos_mes_anterior=None):
    """Carga plantilla, rellena, devuelve los bytes del Excel resultante."""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        shutil.copy(plantilla_path, tmp.name)
        tmp_path = tmp.name

    wb = load_workbook(tmp_path)
    ws = wb['INPUTS']

    ws['C6'] = datos['nombre_cliente']
    ws['F6'] = datos['mes_nombre']
    ws['F7'] = datos['año']
    ws['F8'] = datos['mes_numero']
    ws['C11'] = datos['programados']
    ws['F11'] = datos['completados']
    ws['C12'] = datos['km_totales']
    ws['C13'] = datos['horas_totales']
    ws['F13'] = datos['rodaje_max']
    ws['C19'] = datos['ritmo_medio']

    for i, sem in enumerate(datos['semanas']):
        r = 23 + i
        ws[f'B{r}'] = sem['label']
        ws[f'C{r}'] = sem['km']
        ws[f'E{r}'] = sem['label']
        ws[f'F{r}'] = sem['horas']

    # Sobrescribir _chart_data con valores numéricos directos.
    # Las horas se guardan como número decimal (ej: 2.5 = 2h 30min) con formato
    # "0.0\" h\"" para máxima compatibilidad entre Excel Mac/Windows/LibreOffice.
    # Evitamos los formatos de hora porque openpyxl los convierte a timedelta
    # y Excel Mac no los renderiza bien en gráficos.
    if '_chart_data' in wb.sheetnames:
        ws_cd = wb['_chart_data']
        for i, sem in enumerate(datos['semanas']):
            r = i + 2
            ws_cd[f'A{r}'] = sem['label']
            ws_cd[f'B{r}'] = sem['km']
            # Horas como decimal: 2.5 = 2h 30min
            horas_decimal = sem.get('horas_fraccion_dia', 0) * 24
            ws_cd[f'C{r}'] = round(horas_decimal, 2)
            ws_cd[f'C{r}'].number_format = '0.0" h"'

    # Si tenemos datos del mes anterior, rellenar bloque 3
    if datos_mes_anterior:
        if datos_mes_anterior.get('distancia') is not None:
            ws['C17'] = datos_mes_anterior['distancia']
        if datos_mes_anterior.get('tiempo') is not None:
            ws['F17'] = datos_mes_anterior['tiempo']
        if datos_mes_anterior.get('ritmo') is not None:
            ws['C18'] = datos_mes_anterior['ritmo']
        if datos_mes_anterior.get('rodaje') is not None:
            ws['F18'] = datos_mes_anterior['rodaje']

        # Rellenar fecha de inicio del plan en bloque 1 (C8) si la tenemos
        if datos_mes_anterior.get('fecha_inicio_plan'):
            ws['C8'] = datos_mes_anterior['fecha_inicio_plan']

        # Rellenar distancia objetivo en bloque 1 (C7) si viene del Excel web
        if datos_mes_anterior.get('distancia_objetivo'):
            ws['C7'] = datos_mes_anterior['distancia_objetivo']

    # Guardar a buffer para descarga
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    os.unlink(tmp_path)
    return buffer.getvalue()


# ==========================================================
# INTERFAZ STREAMLIT
# ==========================================================

st.set_page_config(
    page_title="LP Training - Generador de Informes",
    page_icon="🏃",
    layout="centered"
)

# Estilo personalizado con tu paleta amarillo/negro
st.markdown("""
<style>
    .main-header {
        background-color: #1A1A1A;
        color: white;
        padding: 1.5rem;
        border-radius: 8px;
        text-align: center;
        margin-bottom: 2rem;
    }
    .main-header h1 {
        margin: 0;
        font-size: 2rem;
        letter-spacing: 1px;
    }
    .main-header p {
        margin: 0.3rem 0 0 0;
        font-size: 0.9rem;
        opacity: 0.85;
        font-style: italic;
    }
    .stButton > button {
        background-color: #FFD600;
        color: #1A1A1A;
        font-weight: bold;
        border: 2px solid #F9A825;
        padding: 0.5rem 2rem;
    }
    .stButton > button:hover {
        background-color: #F9A825;
        color: #1A1A1A;
        border-color: #1A1A1A;
    }
    .metric-card {
        background-color: #FFF9C4;
        padding: 1rem;
        border-radius: 6px;
        border-left: 4px solid #FFD600;
        margin: 0.5rem 0;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
    <h1>🏃 MÉTODO LP TRAINING</h1>
    <p>Generador de Informes Mensuales · @LP.RTRAINING_</p>
</div>
""", unsafe_allow_html=True)

st.markdown("### Paso 1 · Datos del cliente")

nombre_cliente = st.text_input(
    "Nombre del cliente",
    placeholder="Ej: SERGIO ALONSO",
    help="Se mostrará tal cual en el informe"
).strip().upper()

st.markdown("### Paso 2 · CSV de TrainingPeaks")

archivo_csv = st.file_uploader(
    "Sube el archivo workouts.csv que descargaste de TrainingPeaks",
    type=['csv'],
    help="Exporta desde TrainingPeaks con rango ampliado: desde el lunes anterior al día 1 hasta el domingo posterior al fin de mes. Esto asegura que las semanas que cruzan meses se calculan completas."
)

st.markdown("### Paso 3 · Excel del mes anterior (opcional)")

archivo_anterior = st.file_uploader(
    "Sube el Excel del cliente del mes pasado para rellenar el bloque de comparación automáticamente",
    type=['xlsx'],
    help="Funciona tanto con tu Excel técnico antiguo como con un Excel generado por esta misma web el mes pasado"
)

st.markdown("### Paso 4 · Generar")

generar = st.button("🚀 Generar Excel del informe", disabled=(not nombre_cliente or not archivo_csv))

if generar:
    try:
        with st.spinner("Procesando datos..."):
            df = cargar_csv(archivo_csv)
            year, month = detectar_mes(df)
            mes_str = nombre_mes(month)

            # Primero: calcular semanas asignadas al mes (lun-dom, contienen días del mes)
            semanas, dias_faltantes = calcular_por_semana(df, year, month)

            # Filtrar el df a los días naturales del mes para los KPIs globales
            df_mes = filtrar_df_al_mes(df, year, month)

            # Calcular métricas globales sobre el df filtrado
            programados, completados = calcular_adherencia(df_mes)
            km_totales, rodaje_max, ritmo_medio = calcular_metricas_running(df_mes)
            horas_totales = calcular_horas_totales(df_mes)

            # Procesar mes anterior si se subió
            datos_mes_anterior = None
            if archivo_anterior is not None:
                datos_mes_anterior = extraer_mes_anterior(archivo_anterior)

            datos = {
                'nombre_cliente': nombre_cliente,
                'mes_nombre': mes_str,
                'año': year,
                'mes_numero': month,
                'programados': programados,
                'completados': completados,
                'km_totales': km_totales,
                'horas_totales': horas_totales,
                'rodaje_max': rodaje_max,
                'ritmo_medio': ritmo_medio,
                'semanas': semanas,
            }

            excel_bytes = rellenar_plantilla(PLANTILLA_PATH, datos, datos_mes_anterior)

        # Mostrar resumen
        st.success(f"✅ Informe de **{nombre_cliente}** generado para **{mes_str} {year}**")

        # Aviso si faltan días en el CSV
        # Solo avisamos si los días faltantes son significativos (al inicio o final
        # del rango de semanas), no si son huecos puntuales (que son descansos reales).
        if dias_faltantes:
            primer_falta = min(dias_faltantes)
            ultimo_falta = max(dias_faltantes)
            num_faltas = len(dias_faltantes)
            rango_str = primer_falta.strftime('%d/%m') if num_faltas == 1 else f"{primer_falta.strftime('%d/%m')} al {ultimo_falta.strftime('%d/%m')}"
            st.warning(
                f"⚠️ El CSV no cubre completamente todas las semanas del informe: "
                f"falta{'n' if num_faltas > 1 else ''} {num_faltas} día{'s' if num_faltas > 1 else ''} "
                f"({rango_str}). "
                f"Si esos días el cliente tenía entrenamientos que no aparecen, "
                f"vuelve a TrainingPeaks y exporta con un rango más amplio. "
                f"Si simplemente eran días de descanso, ignora este aviso."
            )

        st.markdown("#### Resumen calculado")

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Adherencia", f"{int(completados/programados*100)}%",
                     f"{completados}/{programados}")
        with col2:
            st.metric("KM totales", f"{km_totales}")
        with col3:
            st.metric("Horas", horas_totales)
        with col4:
            st.metric("Rodaje máximo", f"{rodaje_max} km")

        st.markdown("**Por semana:**")
        df_semanas = pd.DataFrame([{
            'Semana': s['label'],
            'KM': s['km'],
            'Horas': s['horas']
        } for s in semanas])
        st.dataframe(df_semanas, hide_index=True, use_container_width=True)

        # Si se subió Excel del mes anterior, mostrar qué se ha extraído
        if datos_mes_anterior:
            st.markdown("**Datos del mes anterior (extraídos automáticamente):**")
            tipo = "Excel generado por esta web" if datos_mes_anterior['formato'] == 'web' else "Excel técnico antiguo"
            st.caption(f"Formato detectado: {tipo}")

            cma1, cma2, cma3, cma4 = st.columns(4)
            with cma1:
                st.metric("Distancia anterior", f"{datos_mes_anterior['distancia']} km" if datos_mes_anterior['distancia'] else "—")
            with cma2:
                st.metric("Tiempo anterior", datos_mes_anterior['tiempo'] or "—")
            with cma3:
                st.metric("Ritmo anterior", datos_mes_anterior['ritmo'] or "—")
            with cma4:
                st.metric("Rodaje anterior", f"{datos_mes_anterior['rodaje']} km" if datos_mes_anterior['rodaje'] else "—")
        elif archivo_anterior is not None:
            st.warning("⚠️ Subiste un Excel del mes anterior pero no he podido extraer los datos. Revisa que sea el formato correcto. El bloque 3 quedará para rellenar a mano.")

        # Botón de descarga
        nombre_limpio = nombre_cliente.replace(' ', '_')
        nombre_fichero = f"Informe_{nombre_limpio}_{mes_str}_{year}.xlsx"

        st.download_button(
            label="📥 Descargar Excel",
            data=excel_bytes,
            file_name=nombre_fichero,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )

        with st.expander("¿Qué tienes que rellenar manualmente al abrir el Excel?"):
            bloque_3_text = "✅ Rellenado automáticamente desde el Excel del mes anterior" if datos_mes_anterior else "**Bloque 3**: Datos del mes anterior (para la comparación)"
            st.markdown(f"""
            La hoja **INPUTS** del Excel ya tiene rellenado todo lo que se puede sacar del CSV.
            Antes de exportar a PDF, completa también:

            - **Bloque 1**: Distancia objetivo del cliente y fecha de inicio del plan
            - **Bloque 2**: Desnivel acumulado del mes (no viene en el CSV)
            - {bloque_3_text}
            - **Bloques 5 y 6**: Zonas de velocidad y FC (cópialas desde el Panel de Control de TrainingPeaks)
            - **Bloque 7**: Comentarios del entrenador y próximos pasos
            - **Bloque 8**: Histórico anual (actualiza con el km del mes actual)

            Después, exporta a PDF seleccionando solo las hojas «INFORME MENSUAL» y «RESUMEN ANUAL».
            """)

    except Exception as e:
        st.error(f"❌ Hubo un error al procesar el archivo:\n\n```\n{e}\n```")
        st.info("Comprueba que el CSV es el original de TrainingPeaks (sin haberlo abierto y guardado en Excel) y vuelve a intentarlo.")

# Footer
st.markdown("---")
st.markdown(
    "<p style='text-align:center; color:#888; font-size:0.85rem;'>"
    "Lorenzo Pla · Entrenador de corredores online"
    "</p>",
    unsafe_allow_html=True
)
